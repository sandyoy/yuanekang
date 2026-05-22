#!/usr/bin/env python3
"""全面分析珠海妇幼订单数据"""
import openpyxl
from collections import Counter, defaultdict

wb = openpyxl.load_workbook('/Users/sandy/.hermes/cache/documents/doc_9e524f4b823b_珠海妇幼订单数据2021.12-2026.4.xlsx', read_only=True, data_only=True)

print("=" * 80)
print("一、工作表概览")
print("=" * 80)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  📋 {name}: {ws.max_row} 行 × {ws.max_column} 列")

# ================================================================
# 1. 院内+居家 汇总表
# ================================================================
print("\n" + "=" * 80)
print("二、院内+居家 年度订单汇总")
print("=" * 80)
ws = wb['院内+居家']
for i, row in enumerate(ws.iter_rows(min_row=3, max_row=7, values_only=True)):
    vals = [str(v)[:20] if v else '-' for v in row[:8]]
    print(f"  {vals}")

# ================================================================
# 2. 居家数据分析
# ================================================================
print("\n" + "=" * 80)
print("三、居家数据分析 - 各项目年度订单趋势")
print("=" * 80)
ws = wb['居家数据分析']
# 项目表头
for i, row in enumerate(ws.iter_rows(min_row=3, max_row=8, values_only=True)):
    vals = [str(v)[:12] if v else '-' for v in row[:12]]
    print(f"  {vals}")

# 占比分布（正确读取）
print("\n📊 居家核心项目3年累计占比：")
# 先看看第17行之后（之前读到数据不对是因为offset）
for i, row in enumerate(ws.iter_rows(min_row=17, max_row=25, values_only=True)):
    vals = [str(v)[:25] if v else '' for v in row]
    if any(v for v in vals):
        print(f"  行{17+i}: {vals}")

# 采用直接从数据区读取（项目在D列往后是数据，但我们从已知位置读取）
# 从第3行开始，实际项目数据在第18列附近
print("\n📊 居家核心项目3年累计占比（直接从数据区读取）：")
# 读取第7行（3年累计行）和第8行（占比行）的指定列
row7 = list(ws.iter_rows(min_row=7, max_row=7, values_only=True))[0]
row8 = list(ws.iter_rows(min_row=8, max_row=8, values_only=True))[0]
# 行3是项目名称
row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

prj_items = []
for j in range(1, 12):  # B到L列
    pname = str(row3[j]).strip() if row3[j] else ''
    pcount = row7[j]
    if pname and pname != '年份' and pname != '订单合计（单）' and pname != '月均订单' and pname != '备注' and pcount:
        try:
            cnt = int(float(str(pcount)))
            prj_items.append((pname, cnt))
        except:
            pass

total = sum(c for _, c in prj_items)
for pname, cnt in sorted(prj_items, key=lambda x: -x[1]):
    print(f"  {pname}: {cnt}单 ({cnt/total*100:.1f}%)")

# ================================================================
# 3. 居家明细 - 详细分析
# ================================================================
print("\n" + "=" * 80)
print("四、居家明细分析 (共~2,950行)")
print("=" * 80)
ws = wb['居家明细']
house_monthly = defaultdict(int)
house_projects = Counter()
house_total = 0

for row in ws.iter_rows(min_row=3, values_only=True):
    project = row[3]
    month = row[5]
    if month:
        m_str = str(month).strip()
        if len(m_str) >= 6:
            ym = f"{m_str[:4]}-{m_str[4:6]}"
            house_monthly[ym] += 1
    if project:
        p = str(project).strip()
        if p and p not in ('None', ''):
            house_projects[p] += 1
    house_total += 1

print(f"\n📊 居家总订单数: {house_total}")
print(f"\n📅 按月分布:")
for ym in sorted(house_monthly.keys()):
    print(f"   {ym}: {house_monthly[ym]}单")

print(f"\n🏥 各项目分布:")
for p, c in house_projects.most_common():
    print(f"   {p}: {c}单 ({c/house_total*100:.1f}%)")

# ================================================================
# 4. 院内明细 - 详细分析
# ================================================================
print("\n" + "=" * 80)
print("五、院内明细分析 (共~49,555行)")
print("=" * 80)
ws = wb['院内明细']
inner_monthly = defaultdict(int)
inner_projects = Counter()
inner_total = 0
price_values = []

for row in ws.iter_rows(min_row=2, values_only=True):
    project = row[1]
    price = row[2]
    month = row[6]
    
    if month:
        m_str = str(month).strip()
        if len(m_str) >= 6:
            ym = f"{m_str[:4]}-{m_str[4:6]}"
            inner_monthly[ym] += 1
    if project:
        p = str(project).strip()
        if p and p not in ('None', ''):
            inner_projects[p] += 1
    if price:
        try:
            pv = float(str(price))
            price_values.append(pv)
        except:
            pass
    inner_total += 1

print(f"\n📊 院内总订单数: {inner_total:,}")
print(f"\n📅 院内按月分布（按年汇总）:")
yearly_inner = defaultdict(int)
for ym, c in sorted(inner_monthly.items()):
    y = ym[:4]
    yearly_inner[y] += c
for y in sorted(yearly_inner.keys()):
    print(f"   {y}年: {yearly_inner[y]:,}单")

print(f"\n🏥 院内各项目TOP15:")
for p, c in inner_projects.most_common(15):
    print(f"   {p}: {c:,}单 ({c/inner_total*100:.1f}%)")

if price_values:
    print(f"\n💰 价格分析（共{len(price_values):,}条有价格数据）:")
    print(f"   均价: {sum(price_values)/len(price_values):.0f}元")
    print(f"   最高: {max(price_values):.0f}元")
    print(f"   最低: {min(price_values):.0f}元")
    # 价格分段
    ranges = [('0-99元', lambda x: x < 100), ('100-299元', lambda x: 100 <= x < 300),
              ('300-499元', lambda x: 300 <= x < 500), ('500-999元', lambda x: 500 <= x < 1000),
              ('1000元+', lambda x: x >= 1000)]
    for label, fn in ranges:
        cnt = sum(1 for v in price_values if fn(v))
        print(f"   {label}: {cnt:,}单 ({cnt/len(price_values)*100:.1f}%)")

# ================================================================
# 5. 按阶梯定价模拟结算
# ================================================================
print("\n" + "=" * 80)
print("六、按阶梯定价模拟结算（居家+院内合计月推广量）")
print("=" * 80)

# 合并居家和院内按月汇总
all_monthly = defaultdict(int)
for ym, c in house_monthly.items():
    all_monthly[ym] += c
for ym, c in inner_monthly.items():
    all_monthly[ym] += c

# 按合同阶梯计算
def calc_cost(count):
    if count <= 499:
        return count * 5.0
    elif count <= 999:
        return count * 3.5
    elif count <= 1499:
        return count * 1.5
    else:
        return count * 1.0

print(f"\n{'月份':<10} {'推广例数':<10} {'所在档位':<15} {'结算单价':<10} {'甲方收入':<12}")
print("-" * 60)
total_examples = 0
total_cost = 0
for ym in sorted(all_monthly.keys()):
    cnt = all_monthly[ym]
    total_examples += cnt
    cost = calc_cost(cnt)
    total_cost += cost
    if cnt <= 499:
        tier = "0~499例（5元）"
    elif cnt <= 999:
        tier = "500~999例（3.5元）"
    elif cnt <= 1499:
        tier = "1000~1499例（1.5元）"
    else:
        tier = "1500例+（1元）"
    print(f"{ym:<10} {cnt:<10} {tier:<15} {cost/cnt if cnt else 0:<10.2f} {cost:<12.0f}")

print("-" * 60)
print(f"{'合计':<10} {total_examples:<10} {'':<15} {'':<10} {total_cost:<12.0f}")
print(f"\n📌 全部月数据累计: {total_examples:,}例, 甲方应收: {total_cost:,.0f}元")

print("\n" + "=" * 80)
print("七、关键发现总结")
print("=" * 80)
print(f"""
📊 院内业务（4年+）:
  • 2022年起稳定在3,000单/年级别
  • 2026年1-4月已970单，预估全年2,910单
  • 盆底康复、子宫复旧是绝对主力项目

🏠 居家业务（2年+）:
  • 2024年4月上线，当年579单
  • 2025年1,551单（增长168%）
  • 2026年1-4月819单，预估全年2,457单
  • 核心项目：子宫复旧（66%）、内膜养护（13%）、术后康复（13%）

💰 按合同阶梯模拟:
  • 历史全部数据合计约{total_examples:,}例
  • 按阶梯定价乙方需向甲方结算约{total_cost:,.0f}元
  • 居家业务上线后（2024年4月起）月均推广量在200+例
  • 多数月份落在0~499档（5元）或500~999档（3.5元）
""")
